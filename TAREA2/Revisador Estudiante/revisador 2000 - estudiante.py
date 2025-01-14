import subprocess
import os

import difflib
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment

from openpyxl.styles import Font
import socket
import hashlib
import platform
from datetime import datetime

# archivo de salida
nombre = "Resultados_Estudiante.xlsx"
time= 5 # segundos máximos de ejecución de cada test

dict_errores = {
'Semántica':"    El código presentado está correctamente escrito en Python.\n    Pero tiene problemas con el planteamiento del ejercicio y el cómo resolverlo.\n    (Esto puede deberse a:: \n        El formato requerido en el ejercicio no se ha respetado completamente.\n        El código está incompleto.\n        La solución no corresponde al ejercicio solicitado.\n    )",
'IndexError':"    Acceso a posiciones inexistentes en una colección (Listas, Strings).\n    Ejemplo:\nacceder al quinto elemento de una lista que solo tiene 2 elementos.",
'TypeError': "    Uso de una operación con un tipo de dato incompatible.    Ejemplo (sumar un número y un texto):\n    5 + 'a' ",
'NameError':"    Uso de elementos que no se encuentran definidos previamente en el código.\n    Ejemplo:\n    Usar la variable hola antes de asignarle un valor.",
'ValueError' : "    Se ha utilizado un valor inesperado para la operación.\n    Ejemplo:\n    Transformar una palabra en un número.\n    int('ABC') ",
'AttributeError': "    Confusión de las propiedades de los tipos de datos o uso de una propiedad inexistente para un tipo de dato.\n    Ejemplo:\n    Uso de elementos propios de los string en una lista.\n    lista.split(';')",
'SyntaxError' : "    El código no está utilizando correctamente la escritura utilizada en Python\n    Ejemplo:\n    (25 * 3) +  8)\n    (falta/sobra un paréntesis)",
'EOFError':"    Incompatibilidad entre la cantidad de datos que necesita el programa y los entregados.\n    Ejemplo:\n    Pedir más/menos entradas de las que el programa luego utiliza.",
'IndentationError':"    El programa no respeta las tabulaciones (sangría) que necesita la sintaxis de Python.\n    Ejemplo:\n    No escribir más a la derecha luego de un if.",
'Ciclo infinito': "    Existe algún ciclo que tarda más de lo esperado en terminar. Probablemente debido a una tautología.\n    Ejemplo:\n    No actualizar la condición de un ciclo while para que en algún punto se vuelva falsa.",
'Success': "    Test superado con éxito."
}                

def escribir_multi(name,df_list,hoja_list):
    # Escribir el DataFrame en un nuevo archivo Excel
    try:
        with pd.ExcelWriter(nombre, engine='openpyxl') as writer:
            for i in range(len(df_list)):
                df = df_list[i]
                hoja = hoja_list[i]
                df.to_excel(writer, index=False, sheet_name=hoja)
                #writer.save()

        # Cargar el libro de trabajo de Excel
        wb = load_workbook(nombre)

        # Obtener la hoja de cálculo
        hoja = wb["Test"]

        # Configurar la opción wrap_text en True para todas las celdas
        for fila in hoja.iter_rows():
            for celda in fila:
                celda.alignment = Alignment(wrapText=True)

        # Guardar los cambios
        wb.save(name)
    except:
        print("\nEl archivo Resultados_Estudiante.xlsx está abierto. Por favor, ciérralo antes de continuar.")


# Entradas y Salidas
try :
    with open("IO_Esperado.txt", encoding="utf-8") as a:
        lista = a.readlines()
except:
    with open("IO_Esperado.txt") as a:
        lista = a.readlines()
        
testE = []
testS = []
io = "###ENTRADA###"
data = []
for linea in lista:
    if "###ENTRADA###" == linea.strip():
        testS.append(data)
        data = []
        io = "###ENTRADA###"
    elif "###SALIDA###" == linea.strip():
        testE.append(data)
        data = []
        io = "###SALIDA###"        
    else:
        data.append(linea.strip())
testS.append(data)
testS.pop(0)
data = []

# archivos a analizar
alumnos = os.listdir('./estudiantes')
cantidad = str(len(alumnos))

df_list = []

list_test = []
for i in range(1,len(testE)+1):
    list_test.append("TEST "+"0"*(2-len(str(i)))+str(i))
    list_test.append("Entrada "+str(i))
    list_test.append("Salida Estudiante "+str(i))
    list_test.append("Salida Esperada "+str(i))
    list_test.append("Situación Test "+str(i))
list_test.append("Correctas")


#Aquí empieza
actual = 1
correctos = []
fallidos = []
for alumno in alumnos:

    lista_linea = []

    
    if alumno[-3]+alumno[-2]+alumno[-1] == ".py":
        formato = True
    else:
        formato = False
       

    if not formato:
        print("Skip de archivo por formato incorrecto (debe ser un .py)")

        for e in range(len(testE)):
            lista_linea.append(0)
            lista_linea.append("")
            lista_linea.append("")
            lista_linea.append("")
        lista_linea.append(0)
        
    else:
        code = ''
        try:
            with open('./estudiantes/'+alumno,'r', encoding="cp1252") as file:
                code = file.read()
        except UnicodeDecodeError:
            with open('./estudiantes/'+alumno,'r', encoding="utf8") as file:
                code = file.read()
        
        name = socket.gethostname()
        add = socket.gethostbyname(name)
        code_out = '######### '+alumno 
        code_out += '\n######### '+ name
        code_out += '\n######### '+ add
        code_out += '\n######### '+ str(platform.system())
        code_out += '\n######### '+ str(platform.version())
        code_out += '\n######### '+ str(platform.architecture())
        code_out += '\n######### '+ str(platform.node())
        code_out += '\n######### '+ str(platform.processor())
        #code_out += '\n######### '+ str(datetime.now())
        code_out += "\n"+ code
        code_out += "\n#########"
        code = code_out.encode('utf-8')
        hh = hashlib.sha256(code).hexdigest()
        code_out += "\n"+hh        

        with open("log.out",'a') as file:
             file.write("\n\n"+code_out)
        

        
        correctas = 0
        i = 0
        info_test = []
        while i < len(testE):
            
            try:
                
                proceso = subprocess.Popen(
                    ['python', './estudiantes/'+alumno],
                    stdin=subprocess.PIPE,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.PIPE,
                    text=True
                    )
                entrada = "\n".join(testE[i])
                salida, error = proceso.communicate(entrada,timeout=time)
                
            except subprocess.TimeoutExpired:
                proceso.terminate()
                salida, error = "", "Tiempo de ejecución excedido ("+str(time)+"s)"
            salida_esperada = "\n".join(testS[i])

            #print(error)
            if error:
                #print(f"    Se produjo un error: {error}")
                lista_linea.append("0")
                lista_linea.append(entrada)
                lista_linea.append(error)
                lista_linea.append(salida_esperada)
                fallidos.append(str(i+1))
                if "Tiempo de ejecución excedido" in error or "stack overflow" in error:
                    info_test.append("Ciclo infinito")
                    lista_linea.append("Ciclo infinito: "+dict_errores["Ciclo infinito"])
                else:
                    info_test.append(error.split("\n")[-2].split(":")[0])
                    type_error = error.split("\n")[-2].split(":")[0]
                    situacion_error = type_error+": "+dict_errores[type_error]
                    lista_linea.append(situacion_error)
                    

            else:
                resultado = salida.strip()

                if resultado == salida_esperada:
                    lista_linea.append("1")
                    correctas +=1
                    info_test.append("Success")
                    correctos.append(str(i+1))
                    situacion = "Correcta"

                else:
                    lista_linea.append("0")
                    info_test.append("Semántica")
                    fallidos.append(str(i+1))

                    diferencias = difflib.unified_diff(
                        salida_esperada.splitlines(), 
                        resultado.splitlines(), 
                        fromfile='salida_esperada', 
                        tofile='resultado', 
                        lineterm=''
                    )
                    situacion = []
                    num_ln = 0
                    ln_comp = 0
                    list_resultado = resultado
                    list_salida_esperada = salida_esperada
                    for ln in diferencias:
                        if num_ln > 2:
                            situacion.append(ln) 
                        num_ln += 1
                        
                    situacion.append('''
#############
- Indica que una línea está en la salida esperada, pero no en su resultado.
+ Indica que una línea está en su resultado, pero no en salida esperada.''')
                    
                    ln_comp = 0
                    list_resultado = resultado.split("\n")
                    list_salida_esperada = salida_esperada.split("\n")
                    dice = ""
                    decir = ""
                    ln_resultado = []
                    ln_esperada = []
                    diff_flag = False
                    for ln in range(min(len(list_resultado),len(list_salida_esperada))):
                        if list_resultado[ln] != list_salida_esperada[ln]:
                            ln_resultado =  list_resultado[ln].split(" ")
                            ln_esperada =  list_salida_esperada[ln].split(" ")
                            for pal in range( min(len(ln_resultado),len(ln_esperada)) ):
                                if ln_resultado[pal] != ln_esperada[pal]:
                                    dice = ln_resultado[pal]
                                    decir = ln_esperada[pal]
                                    diff_flag = True
                                    break
                            if diff_flag:
                                break
                                    
                    situacion.append( "#############\n")
                    if len(ln_resultado) < len(ln_esperada):
                        situacion.append("Existe un espacio extra.")
                    elif len(ln_resultado) > len(ln_esperada):
                        situacion.append("El formato de espaciado no corresponde.")
                    else:
                        situacion.append( "Dice: "+ dice )            
                        situacion.append( "Debería decir: "+ decir )            
                    situacion = "\n".join(situacion)

                lista_linea.append(entrada)
                lista_linea.append(resultado)
                lista_linea.append(salida_esperada)
                lista_linea.append(situacion)

            i+=1

        lista_linea.append(str(correctas))
        

    try:
        print(str(actual)+"/"+cantidad +" "+alumno+": "+str(correctas)+"/"+str(len(testS)))
        print()
        for e in range(len(info_test)):
        
            if info_test[e] == "Success":
                print("Test "+"0"*(2-len(str(e+1)))+str(e+1)+": "+dict_errores[info_test[e]])
            else :
                print("\nTest "+"0"*(2-len(str(e+1)))+str(e+1)+": Error de "+info_test[e])
                print(dict_errores[info_test[e]])

        print()
        df_list.append(lista_linea)
        actual += 1
    
    except:
        print("El archivo no existe o tiene formato incorrecto.")
    
        
df = pd.DataFrame(df_list , columns = list_test)
data_frames_lista = [df]
hojas_lista = ['Test']
    
escribir_multi(nombre,data_frames_lista,hojas_lista)

print("\n\n>>>>>>   COMPLETADO...   <<<<<<<\n\n")
print("    Test correctos: " + " ".join(correctos))
print("    Test fallidos: " + " ".join(fallidos))

